import base64, io, json, os, platform, socket, subprocess, time, webbrowser, logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
from urllib.parse import urlparse
import requests
from dotenv import load_dotenv, set_key
from PIL import ImageGrab

ENV_PATH=Path(__file__).with_name('.env');load_dotenv(ENV_PATH,override=True)
BASE=os.environ.get('JARVIS_URL','').strip().rstrip('/')
PAIRING_SECRET=os.environ.get('JARVIS_PAIRING_SECRET','').strip()
TOKEN=os.environ.get('JARVIS_DEVICE_TOKEN','').strip()
DEVICE_NAME=os.environ.get('DEVICE_NAME',socket.gethostname()).strip() or socket.gethostname()
POLL=max(1,int(os.environ.get('POLL_SECONDS','2')))
INDEX_INTERVAL=max(120,int(os.environ.get('INDEX_INTERVAL_SECONDS','600')))
MAX_INDEX_FILES=max(100,min(5000,int(os.environ.get('MAX_INDEX_FILES','1200'))))
ENABLE_DESKTOP_AUTOMATION=os.environ.get('ENABLE_DESKTOP_AUTOMATION','false').lower()=='true'

LOG_DIR=Path(__file__).with_name('logs');LOG_DIR.mkdir(exist_ok=True);LOG_FILE=LOG_DIR/'agent.log'
logger=logging.getLogger('jarvis-agent');logger.setLevel(logging.INFO)
if not logger.handlers:
    fh=RotatingFileHandler(LOG_FILE,maxBytes=2_000_000,backupCount=4,encoding='utf-8');fh.setFormatter(logging.Formatter('%(asctime)s | %(levelname)s | %(message)s'));logger.addHandler(fh)
    sh=logging.StreamHandler();sh.setFormatter(logging.Formatter('%(message)s'));logger.addHandler(sh)

ALLOWED_APPS={
 'chrome':[r'C:\Program Files\Google\Chrome\Application\chrome.exe',r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'],
 'edge':[r'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe',r'C:\Program Files\Microsoft\Edge\Application\msedge.exe'],
 'notepad':['notepad.exe'],'explorer':['explorer.exe'],'calculator':['calc.exe'],'paint':['mspaint.exe']
}
TEXT_EXT={'.txt','.md','.csv','.json','.log','.ini','.yaml','.yml','.xml','.html','.htm','.py','.js','.ts','.tsx','.jsx','.css','.sql'}
DOC_EXT={'.pdf','.docx','.xlsx'}

def headers(): return {'Authorization':f'Bearer {TOKEN}','Content-Type':'application/json'}
def validate_base_url():
    if not BASE.startswith(('http://','https://')): raise SystemExit('HATA: JARVIS_URL http:// veya https:// ile başlamalı.')
def save_token(token):
    if not ENV_PATH.exists(): ENV_PATH.write_text('',encoding='utf-8')
    set_key(str(ENV_PATH),'JARVIS_DEVICE_TOKEN',token)
def pair_if_needed():
    global TOKEN
    if TOKEN:return
    if not PAIRING_SECRET:raise SystemExit('HATA: İlk eşleştirme için JARVIS_PAIRING_SECRET gerekli.')
    logger.info(f'JARVIS: {DEVICE_NAME} eşleştiriliyor...')
    r=requests.post(BASE+'/api/agent/pair',json={'pairing_secret':PAIRING_SECRET,'name':DEVICE_NAME},timeout=20)
    if not r.ok:raise SystemExit(f'Eşleştirme başarısız ({r.status_code}): {r.text}')
    TOKEN=r.json()['device_token'];save_token(TOKEN);logger.info('JARVIS: cihaz başarıyla eşleştirildi.')

def safe_path(target, must_exist=True):
    p=Path(os.path.expandvars(os.path.expanduser(target))).resolve()
    roots=[Path.home().resolve(),Path('C:/Users/Public').resolve()]
    if not any(str(p).lower()==str(r).lower() or str(p).lower().startswith(str(r).lower()+os.sep.lower()) for r in roots):raise PermissionError('Bu yol güvenli kullanıcı alanının dışında.')
    if must_exist and not p.exists():raise FileNotFoundError(str(p))
    return p

def open_app(key):
    key=key.strip().lower(); candidates=ALLOWED_APPS.get(key)
    if not candidates:raise ValueError(f'Uygulama izin listesinde değil: {key}')
    chosen=None
    for c in candidates:
        if ('\\' in c or '/' in c) and c.lower().endswith('.exe'):
            if Path(c).exists():chosen=c;break
        else:chosen=c;break
    if not chosen:raise FileNotFoundError(f'{key} bulunamadı.')
    subprocess.Popen([chosen],shell=False);return f'Uygulama açıldı: {key}'

def screenshot_data_url():
    img=ImageGrab.grab(all_screens=True);buf=io.BytesIO();img.save(buf,format='JPEG',quality=72,optimize=True)
    return 'data:image/jpeg;base64,'+base64.b64encode(buf.getvalue()).decode('ascii')

def clipboard_read():
    try:
        import tkinter as tk
        root=tk.Tk();root.withdraw();txt=root.clipboard_get();root.destroy();return str(txt)[:12000]
    except Exception as e: return f'Panoya erişilemedi: {e}'

def execute(action,target,payload=None):
    payload=payload or {}
    if action=='open_url':
        raw=target.strip();raw=raw if raw.startswith(('https://','http://')) else 'https://'+raw;u=urlparse(raw)
        if u.scheme not in ('http','https') or not u.netloc:raise ValueError('Geçersiz URL.')
        webbrowser.open(raw);return f'URL açıldı: {raw}'
    if action=='open_app':return open_app(target)
    if action=='open_folder':
        p=safe_path(target); 
        if not p.is_dir():raise ValueError('Klasör değil.')
        subprocess.Popen(['explorer.exe',str(p)],shell=False);return f'Klasör açıldı: {p}'
    if action=='open_file':
        p=safe_path(target)
        if not p.is_file():raise ValueError('Dosya değil.')
        os.startfile(str(p));return f'Dosya açıldı: {p}'
    if action=='system_info':return json.dumps({'computer':socket.gethostname(),'os':platform.platform(),'python':platform.python_version(),'user_home':str(Path.home())},ensure_ascii=False)
    if action=='screenshot':return screenshot_data_url()
    if action=='clipboard_read':return clipboard_read()
    if action=='index_files':
        n=sync_file_index();return f'{n} dosya indekslendi.'
    if action=='desktop_click':
        if not ENABLE_DESKTOP_AUTOMATION:raise PermissionError('Masaüstü otomasyonu kapalı.')
        import pyautogui
        parts=[x.strip() for x in target.split(',')];
        if len(parts)!=2:raise ValueError('Koordinat x,y olmalı.')
        x,y=int(parts[0]),int(parts[1]);w,h=pyautogui.size()
        if not (0<=x<w and 0<=y<h):raise ValueError('Koordinat ekran dışında.')
        pyautogui.click(x,y);return f'Tıklama yapıldı: {x},{y}'
    if action=='desktop_type':
        if not ENABLE_DESKTOP_AUTOMATION:raise PermissionError('Masaüstü otomasyonu kapalı.')
        import pyautogui
        txt=str(target)[:1000];pyautogui.write(txt,interval=0.015);return 'Metin yazıldı.'
    raise ValueError('Desteklenmeyen işlem')

def extract_excerpt(path):
    ext=path.suffix.lower()
    try:
        if ext in TEXT_EXT:
            return path.read_text(encoding='utf-8',errors='ignore')[:10000]
        if ext=='.pdf':
            from pypdf import PdfReader
            r=PdfReader(str(path));return '\n'.join((p.extract_text() or '') for p in r.pages[:8])[:10000]
        if ext=='.docx':
            from docx import Document
            d=Document(str(path));return '\n'.join(p.text for p in d.paragraphs[:300])[:10000]
        if ext=='.xlsx':
            from openpyxl import load_workbook
            wb=load_workbook(str(path),read_only=True,data_only=True);out=[]
            for ws in wb.worksheets[:3]:
                for row in ws.iter_rows(max_row=150,values_only=True):out.append(' | '.join('' if v is None else str(v) for v in row[:20]))
            wb.close();return '\n'.join(out)[:10000]
    except Exception:return ''
    return ''

def index_roots():
    raw=os.environ.get('INDEX_ROOTS','%USERPROFILE%\\Documents;%USERPROFILE%\\Desktop;%USERPROFILE%\\Downloads')
    roots=[]
    for x in raw.split(';'):
        if not x.strip():continue
        try:
            p=safe_path(os.path.expandvars(x.strip()))
            if p.is_dir():roots.append(p)
        except Exception:pass
    return roots

def collect_files():
    rows=[]
    for root in index_roots():
        for path in root.rglob('*'):
            if len(rows)>=MAX_INDEX_FILES:return rows
            try:
                if not path.is_file() or path.name.startswith('~$') or path.stat().st_size>25_000_000:continue
                ext=path.suffix.lower()
                if ext not in TEXT_EXT|DOC_EXT:continue
                st=path.stat();excerpt=extract_excerpt(path)
                rows.append({'path':str(path),'name':path.name,'extension':ext,'size':st.st_size,'modified_at':time.strftime('%Y-%m-%dT%H:%M:%SZ',time.gmtime(st.st_mtime)),'excerpt':excerpt[:3000],'searchable_text':excerpt[:10000]})
            except Exception:continue
    return rows

def sync_file_index():
    files=collect_files();total=0
    for i in range(0,len(files),200):
        r=requests.post(BASE+'/api/agent/sync-files',headers=headers(),json={'files':files[i:i+200]},timeout=60);r.raise_for_status();total+=len(files[i:i+200])
    logger.info(f'Dosya indeksi senkronlandı: {total}');return total

def main():
    validate_base_url();pair_if_needed();logger.info(f'JARVIS Windows Agent çevrimiçi. Sunucu: {BASE}')
    last_index=0
    while True:
        try:
            if time.time()-last_index>INDEX_INTERVAL:
                try:sync_file_index()
                except Exception as e:logger.warning(f'İndeksleme: {e}')
                last_index=time.time()
            r=requests.post(BASE+'/api/agent/pull',headers=headers(),json={},timeout=25)
            if r.status_code==401:raise RuntimeError('Cihaz tokenı geçersiz; yeniden eşleştirme gerekli.')
            r.raise_for_status();cmd=r.json().get('command')
            if cmd:
                logger.info(f"Komut: {cmd['action']} -> {str(cmd.get('target',''))[:200]}")
                ok=True
                try:result=execute(cmd['action'],cmd.get('target',''),cmd.get('payload') or {})
                except Exception as e:ok=False;result=str(e)
                rr=requests.post(BASE+'/api/agent/result',headers=headers(),json={'id':cmd['id'],'ok':ok,'result':result},timeout=30);rr.raise_for_status();logger.info(f'Sonuç: {str(result)[:500]}')
            time.sleep(POLL)
        except KeyboardInterrupt:logger.info('JARVIS Agent kapatıldı.');break
        except Exception as e:logger.error(f'Agent: {e}');time.sleep(5)
if __name__=='__main__':main()
